import docx
import openpyxl

ExcelWorkBook = openpyxl.Workbook()
Sheet = ExcelWorkBook.create_sheet(title='Sheet1',index=0)
ROW = 1
COLUMN = 1

def openDoc(docPath):
    docObj = docx.Document(docPath)
    result = []
    paragraphs = docObj.paragraphs
    for p in paragraphs:
        #fullText.append(p.text)
        if processing(p.text) is None:
            continue
        result.extend(processing(p.text))
    if result:
        return result

def processing(text):
    startPoint = None
    #endPoint = None
    word = []
    fullName = []
    for index in range(len(text)):
        if text[index] == "(":
            startPoint = index
            continue
        elif text[index] == "," and startPoint:
            fullName.append(text[startPoint+1:index])
            startPoint = index+1
            continue
        elif text[index] == ")" and startPoint:
            fullName.append(text[startPoint+1:index])
            word.append(fullName)
            startPoint = None
            fullName = []
            continue
    if word:
        return word

def printAdv(listTmp):
    for item in listTmp:
        if isinstance(item,list) and len(item)>2:
            printAdv(item)
        else:
            print(item)

def processingForMe(listTmp):
    for item in listTmp:
        if isinstance(item,list) and len(item)>2:
            printAdv(item)
        else:
            print(item)

def excelProcessing(dataList,ROW):
    for item in dataList:
        if isinstance(item,list) and len(item)>2:
            excelProcessing(item,len(item))
        elif len(item)>1:
            COLUMN = 1
            for subItem in item:
                Sheet.cell(ROW,COLUMN,subItem)
                COLUMN = COLUMN + 1
            ROW = ROW + 1

def saveExcelFile():
    excelFileName = input("Please type the name of excel file")
    ExcelWorkBook.save(excelFileName+".xlsx")


if __name__ == "__main__":
    docPath = input("Please type the path of document"+"\n")
    result = openDoc(docPath)
    printAdv(result)
    excelProcessing(result,1)
    saveExcelFile()